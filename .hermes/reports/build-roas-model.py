# -*- coding: utf-8 -*-
# 环境无 LibreOffice 无法重算公式，按 skill 降级：Python 计算后直接写值
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color='0000FF')
BLACK = Font(color='000000')
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
YELLOW = PatternFill('solid', start_color='FFFF00')
HL = PatternFill('solid', start_color='FFF2CC')
RED_TXT = Font(color='C00000', bold=True)
THIN = Border(left=Side('thin', color='BFBFBF'), right=Side('thin', color='BFBFBF'),
              top=Side('thin', color='BFBFBF'), bottom=Side('thin', color='BFBFBF'))
PCT = '0.0%'
MONEY = '$#,##0'
X = '0.0x'

# ---- 假设（可调）----
AOV = 150.0
MARGIN = 0.50
LOGI = 0.10
PAY = 0.03
ROAS = 2.5

contrib = MARGIN - LOGI - PAY          # 广告前贡献率
ad_share = 1 / ROAS                    # 广告费占营收比
paid_margin = contrib - ad_share       # 付费净利率
be_roas = 1 / contrib                  # 盈亏平衡 ROAS
organic_margin = contrib               # 自然净利率
gap = organic_margin - paid_margin

wb = Workbook()

# ---------- Sheet 1 假设与结论 ----------
ws = wb.active
ws.title = '假设与结论'
ws['A1'] = 'ZprintPro 付费 vs 自然流量 利润测算（每 $100 营收口径）'
ws['A1'].font = TITLE
ws['A2'] = '蓝色=可调输入（改后需重跑模型）；黑色=计算结果；黄色底=待用真实数据校准'
ws['A2'].font = Font(italic=True, size=9, color='808080')

rows = [
    ('客单价 AOV ($)', AOV, MONEY, True, '推演值：C端$30-80/B端$150-500 混合，待真实订单校准', False),
    ('毛利率（扣除生产+材料+人工）', MARGIN, PCT, True, '印刷行业典型 40-60%，待真实成本数据校准', False),
    ('跨境物流占营收比', LOGI, PCT, True, 'DHL 直送，小单占比更高，待校准', False),
    ('支付手续费率（Airwallex）', PAY, PCT, True, '约 2.8-3.4%', False),
    ('广告前贡献率', contrib, PCT, False, '=毛利率-物流-支付费 = 50%-10%-3%', False),
    ('投放 ROAS', ROAS, X, True, 'ROAS=每$1广告费产生的营收', False),
    ('广告费占营收比', ad_share, PCT, False, '=1/ROAS，ROAS 2.5 即占 40%', False),
    ('付费渠道净利润率', paid_margin, PCT, False, '=贡献率-广告费占比', paid_margin < 0),
    ('盈亏平衡 ROAS', be_roas, X, False, 'ROAS 低于此值=每单亏损', False),
    ('自然流量(SEO/GEO)净利润率', organic_margin, PCT, False, '获客成本≈0，贡献率全部落袋', False),
    ('利润率差距（自然-付费）', gap, PCT, False, '每 $100 营收的差距', False),
]
ws.cell(4, 1, '项目').font = BOLD
ws.cell(4, 2, '数值').font = BOLD
ws.cell(4, 3, '说明').font = BOLD
for c in range(1, 4):
    ws.cell(4, c).border = THIN
for i, (label, val, fmt, is_input, note, is_red) in enumerate(rows):
    rr = 5 + i
    ws.cell(rr, 1, label).border = THIN
    cell = ws.cell(rr, 2, val)
    cell.number_format = fmt
    cell.border = THIN
    if is_red:
        cell.font = RED_TXT
        cell.fill = PatternFill('solid', start_color='F8CBAD')
    elif is_input:
        cell.font = BLUE
        cell.fill = YELLOW
    else:
        cell.font = BLACK
    ws.cell(rr, 3, note).font = Font(size=9, color='595959')

ws.cell(17, 1, '结论').font = BOLD
concl = [
    f'① ROAS {ROAS} 时广告费吃掉营收的 {ad_share:.0%}，高于广告前贡献率 {contrib:.0%} → 每单亏 {abs(paid_margin):.0%}，卖得越多亏得越多。',
    f'② 盈亏平衡需要 ROAS≥{be_roas:.1f}；要赚到 10% 净利润率需要 ROAS≥{1/(contrib-0.10):.1f}，印刷电商极少有投放能稳定做到。',
    f'③ 自然流量（SEO/GEO）同样 $100 营收净赚 ${organic_margin*100:.0f}，利润率 {organic_margin:.0%} vs 付费 {paid_margin:.0%}，差距 {gap:.0%}。',
    '④ SEO/GEO 内容资产可长期复用，边际成本趋零；广告费是每单重复支出的沉没成本。',
]
for i, t in enumerate(concl):
    ws.cell(18 + i, 1, t).font = Font(size=10)

ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 54
for rr in range(4, 22):
    ws.row_dimensions[rr].height = 20

# ---------- Sheet 2 ROAS 敏感性矩阵 ----------
ws2 = wb.create_sheet('ROAS敏感性矩阵')
ws2['A1'] = '净利润率敏感性：毛利率 × ROAS（物流 10% + 支付 3% 固定扣除）'
ws2['A1'].font = TITLE
ws2['A2'] = '红色=该组合亏损；黄色列=当前 ROAS 2.5；最右列=该毛利率的盈亏平衡 ROAS'
ws2['A2'].font = Font(italic=True, size=9, color='808080')

roas_list = [1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 5.0]
margins = [0.40, 0.45, 0.50, 0.55, 0.60]
ws2.cell(4, 1, '毛利率 \\ ROAS').font = BOLD
ws2.cell(4, 1).border = THIN
for j, roas in enumerate(roas_list):
    c = ws2.cell(4, 2 + j, roas)
    c.font = BOLD
    c.number_format = X
    c.border = THIN
    c.alignment = Alignment(horizontal='center')
    if roas == 2.5:
        c.fill = YELLOW
be_col = 2 + len(roas_list)
hc = ws2.cell(4, be_col, '盈亏平衡ROAS')
hc.font = BOLD
hc.border = THIN
hc.alignment = Alignment(horizontal='center')
hc.fill = HL
for i, m in enumerate(margins):
    rr = 5 + i
    mc = ws2.cell(rr, 1, m)
    mc.font = BLUE
    mc.number_format = PCT
    mc.border = THIN
    for j, roas in enumerate(roas_list):
        v = (m - LOGI - PAY) - 1 / roas
        cell = ws2.cell(rr, 2 + j, v)
        cell.number_format = PCT
        cell.border = THIN
        cell.alignment = Alignment(horizontal='center')
        if v < 0:
            cell.font = RED_TXT
        else:
            cell.font = BLACK
        if roas == 2.5:
            cell.fill = YELLOW
    be_cell = ws2.cell(rr, be_col, 1 / (m - LOGI - PAY))
    be_cell.number_format = X
    be_cell.font = Font(bold=True)
    be_cell.border = THIN
    be_cell.alignment = Alignment(horizontal='center')
    be_cell.fill = HL
note_r = 5 + len(margins) + 1
ws2.cell(note_r, 1, '注：矩阵值<0 表示该 毛利率×ROAS 组合每单亏损；毛利率需先扣物流+支付才是广告前贡献率。').font = Font(size=9, color='595959')
ws2.column_dimensions['A'].width = 20
for j in range(len(roas_list)):
    ws2.column_dimensions[get_column_letter(2 + j)].width = 11
ws2.column_dimensions[get_column_letter(be_col)].width = 14
for rr in range(4, note_r + 1):
    ws2.row_dimensions[rr].height = 20

# ---------- Sheet 3 每$100营收对比 ----------
ws3 = wb.create_sheet('每$100营收对比')
ws3['A1'] = '同样 $100 营收：付费渠道(ROAS 2.5) vs 自然流量(SEO/GEO)'
ws3['A1'].font = TITLE
ws3.cell(3, 1, '项目').font = BOLD
ws3.cell(3, 2, '付费渠道').font = BOLD
ws3.cell(3, 3, '自然流量').font = BOLD
for c in range(1, 4):
    ws3.cell(3, c).border = THIN

paid_cost = -MARGIN * 100
paid_logi = -LOGI * 100
paid_payfee = -PAY * 100
paid_ad = -100 / ROAS
paid_profit = 100 + paid_cost + paid_logi + paid_payfee + paid_ad
org_profit = 100 + paid_cost + paid_logi + paid_payfee

cmp_rows = [
    ('营收', 100, 100, MONEY, False),
    ('生产成本（按毛利率）', paid_cost, paid_cost, MONEY, False),
    ('跨境物流', paid_logi, paid_logi, MONEY, False),
    ('支付手续费', paid_payfee, paid_payfee, MONEY, False),
    ('广告费（1/ROAS）', paid_ad, 0, MONEY, False),
    ('净利润', paid_profit, org_profit, MONEY, True),
    ('净利润率', paid_profit / 100, org_profit / 100, PCT, True),
]
for i, (label, b, c, fmt, hl) in enumerate(cmp_rows):
    rr = 4 + i
    ws3.cell(rr, 1, label).border = THIN
    for col, val in ((2, b), (3, c)):
        cell = ws3.cell(rr, col, val)
        cell.number_format = fmt
        cell.border = THIN
        cell.alignment = Alignment(horizontal='right')
        cell.font = RED_TXT if (isinstance(val, (int, float)) and val < 0) else BLACK
        if hl:
            cell.font = Font(bold=True, color='C00000' if val < 0 else '006100')
            cell.fill = HL
ws3.cell(12, 1, f'解读：付费渠道每 $100 营收倒亏约 ${abs(paid_profit):.0f}（ROAS {ROAS} 未达盈亏平衡 {be_roas:.1f}）；自然流量净赚约 ${org_profit:.0f}。').font = Font(size=10)
ws3.cell(13, 1, f'差距约 {gap:.0%}——同样的订单，走自然流量是利润，走付费投放是亏损。').font = Font(size=10)
ws3.column_dimensions['A'].width = 26
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 14
for rr in range(3, 14):
    ws3.row_dimensions[rr].height = 20

out = r'F:\zprintpro-nextjs\.hermes\reports\2026-08-10-ROAS利润测算.xlsx'
wb.save(out)
print('saved:', out)
print(f'contrib={contrib:.3f} ad_share={ad_share:.3f} paid_margin={paid_margin:.3f} be_roas={be_roas:.3f} organic={organic_margin:.3f} gap={gap:.3f}')
print(f'paid_profit={paid_profit:.2f} org_profit={org_profit:.2f}')
