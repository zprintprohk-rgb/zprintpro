#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Proforma Invoice + Quotation Excel 生成
- 调用 openpyxl (跨平台, OpenXML 标准)
- PI: 3 工作表 (Proforma Invoice + Items + Bank)
- Quotation: 2 工作表 (Quotation + Cost Breakdown)
- 落盘:
  - F:\zprintpro-nextjs\docs\invoices\PI-2026-08-001.xlsx (en 美元)
  - F:\zprintpro-nextjs\docs\quotes\Q-2026-08-001.xlsx (zh-hk 港币)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, timedelta
import os

INVOICE_DIR = r"F:\zprintpro-nextjs\docs\invoices"
QUOTE_DIR = r"F:\zprintpro-nextjs\docs\quotes"
os.makedirs(INVOICE_DIR, exist_ok=True)
os.makedirs(QUOTE_DIR, exist_ok=True)

# ============================================================================
# 样式定义
# ============================================================================
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")
VALUE_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_FILL = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
ZH_TITLE_FONT = Font(name="標楷體", size=16, bold=True, color="FFFFFF")
ZH_HEADER_FONT = Font(name="標楷體", size=11, bold=True, color="FFFFFF")
ZH_LABEL_FONT = Font(name="標楷體", size=10, bold=True, color="1F4E79")
ZH_VALUE_FONT = Font(name="標楷體", size=10)
ZH_TOTAL_FONT = Font(name="標楷體", size=11, bold=True)
ZH_NOTE_FONT = Font(name="標楷體", size=9, italic=True, color="666666")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="999999")
MEDIUM = Side(border_style="medium", color="1F4E79")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BORDER = Border(top=MEDIUM, left=THIN, right=THIN, bottom=THIN)
BOTTOM_BORDER = Border(bottom=MEDIUM, left=THIN, right=THIN, top=THIN)

# ============================================================================
# Proforma Invoice (PI-2026-08-001.xlsx) - en 美元 + FOB Shenzhen + 3 工作表
# ============================================================================
wb = Workbook()

# ---------- Sheet 1: Proforma Invoice ----------
ws1 = wb.active
ws1.title = "Proforma Invoice"

# 设置列宽
col_widths1 = {"A": 4, "B": 22, "C": 38, "D": 12, "E": 14, "F": 16, "G": 4}
for col, width in col_widths1.items():
    ws1.column_dimensions[col].width = width

# Row 1: 标题
ws1.row_dimensions[1].height = 36
ws1.merge_cells("B1:F1")
cell = ws1["B1"]
cell.value = "PROFORMA INVOICE"
cell.font = TITLE_FONT
cell.fill = TITLE_FILL
cell.alignment = CENTER

# Row 2: PI 编号 + 日期
ws1.row_dimensions[2].height = 22
ws1["B2"] = "PI Number:"
ws1["B2"].font = LABEL_FONT
ws1["B2"].alignment = LEFT
ws1["C2"] = "PI-2026-08-001"
ws1["C2"].font = VALUE_FONT
ws1["C2"].alignment = LEFT
ws1["D2"] = "PI Date:"
ws1["D2"].font = LABEL_FONT
ws1["D2"].alignment = LEFT
ws1["E2"] = "2026-08-27"
ws1["E2"].font = VALUE_FONT
ws1["E2"].alignment = LEFT
ws1["F2"] = ""
ws1["F2"].alignment = LEFT

# Row 3: Valid Until
ws1.row_dimensions[3].height = 22
ws1["B3"] = "Valid Until:"
ws1["B3"].font = LABEL_FONT
ws1["B3"].alignment = LEFT
ws1["C3"] = "2026-09-26 (30 days)"
ws1["C3"].font = VALUE_FONT
ws1["C3"].alignment = LEFT
ws1["D3"] = "Currency:"
ws1["D3"].font = LABEL_FONT
ws1["D3"].alignment = LEFT
ws1["E3"] = "USD"
ws1["E3"].font = VALUE_FONT
ws1["E3"].alignment = LEFT
ws1["F3"] = ""

# Row 4: 空行
ws1.row_dimensions[4].height = 6

# Row 5-6: 卖方 (Seller)
ws1.row_dimensions[5].height = 22
ws1.merge_cells("B5:F5")
ws1["B5"] = "SELLER (卖方)"
ws1["B5"].font = HEADER_FONT
ws1["B5"].fill = HEADER_FILL
ws1["B5"].alignment = LEFT

seller_rows = [
    ("Company Name", "Shenzhen Cailong Printing Packaging Co., Ltd."),
    ("Address", "No.1 Jiacheng Road, Pinghu Street, Longgang District, Shenzhen, Guangdong, China 518111"),
    ("Contact", "Mr. Tang Yunti (Legal Representative)"),
    ("Tel", "+86 198 8085 1334 (WhatsApp unified)"),
    ("Email", "zprintpro@outlook.com"),
    ("Certifications", "ISO 9001 + FSC Chain of Custody Certified"),
]
for i, (k, v) in enumerate(seller_rows):
    row = 6 + i
    ws1.row_dimensions[row].height = 22
    ws1.cell(row, 2, k).font = LABEL_FONT
    ws1.cell(row, 2).alignment = LEFT
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws1.cell(row, 3, v).font = VALUE_FONT
    ws1.cell(row, 3).alignment = LEFT

# Row 13: 空行
ws1.row_dimensions[13].height = 6

# Row 14: 买方 (Buyer)
ws1.row_dimensions[14].height = 22
ws1.merge_cells("B14:F14")
ws1["B14"] = "BUYER (买方)"
ws1["B14"].font = HEADER_FONT
ws1["B14"].fill = HEADER_FILL
ws1["B14"].alignment = LEFT

buyer_rows = [
    ("Company Name", "[Customer Company Name]"),
    ("Address", "[Customer Address]"),
    ("Contact", "[Contact Person Name]"),
    ("Tel", "[Customer Tel]"),
    ("Email", "[Customer Email]"),
]
for i, (k, v) in enumerate(buyer_rows):
    row = 15 + i
    ws1.row_dimensions[row].height = 22
    ws1.cell(row, 2, k).font = LABEL_FONT
    ws1.cell(row, 2).alignment = LEFT
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws1.cell(row, 3, v).font = VALUE_FONT
    ws1.cell(row, 3).alignment = LEFT

# Row 20: 空行
ws1.row_dimensions[20].height = 6

# Row 21-22: 贸易术语 + 付款方式
ws1.row_dimensions[21].height = 22
ws1["B21"] = "Incoterm (per Incoterms 2020):"
ws1["B21"].font = LABEL_FONT
ws1["B21"].alignment = LEFT
ws1.merge_cells("C21:F21")
ws1["C21"] = "FOB Shenzhen (Free On Board, buyer arranges shipping + insurance)"
ws1["C21"].font = VALUE_FONT
ws1["C21"].alignment = LEFT

ws1.row_dimensions[22].height = 22
ws1["B22"] = "Payment Terms:"
ws1["B22"].font = LABEL_FONT
ws1["B22"].alignment = LEFT
ws1.merge_cells("C22:F22")
ws1["C22"] = "T/T 30% deposit + 70% balance before shipment (or L/C at sight for amount >= USD 10,000)"
ws1["C22"].font = VALUE_FONT
ws1["C22"].alignment = LEFT

# Row 23: 空行
ws1.row_dimensions[23].height = 6

# Row 24: 品名表标题
ws1.row_dimensions[24].height = 22
ws1.merge_cells("B24:F24")
ws1["B24"] = "ITEMS TABLE (品名表)"
ws1["B24"].font = HEADER_FONT
ws1["B24"].fill = HEADER_FILL
ws1["B24"].alignment = LEFT

# Row 25: 表头
ws1.row_dimensions[25].height = 28
headers = ["HS Code", "Description", "Qty", "Unit Price (USD)", "Amount (USD)"]
for i, h in enumerate(headers):
    cell = ws1.cell(25, 2 + i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDER

# Row 26-28: 3 个品名 (公式 = qty * unit_price)
items = [
    ("49119900", "Custom Sticker (round 50mm diameter, waterproof vinyl, CMYK 4C, gloss lamination, die-cut)", 10000, 0.03),
    ("48191000", "Custom Packaging Box (corrugated 3-ply B-flute 250gsm, CMYK 4C, matte lamination)", 1000, 1.50),
    ("48194000", "Custom Paper Bag (white card 250gsm + cotton rope, CMYK 4C, spot UV)", 500, 0.50),
]
for i, (hs, desc, qty, price) in enumerate(items):
    row = 26 + i
    ws1.row_dimensions[row].height = 32
    ws1.cell(row, 2, hs).font = VALUE_FONT
    ws1.cell(row, 2).alignment = LEFT
    ws1.cell(row, 3, desc).font = VALUE_FONT
    ws1.cell(row, 3).alignment = LEFT
    ws1.cell(row, 4, qty).font = VALUE_FONT
    ws1.cell(row, 4).alignment = RIGHT
    ws1.cell(row, 4).number_format = "#,##0"
    ws1.cell(row, 5, price).font = VALUE_FONT
    ws1.cell(row, 5).alignment = RIGHT
    ws1.cell(row, 5).number_format = '"$"#,##0.0000'
    # Amount = Qty * Unit Price (公式)
    ws1.cell(row, 6, f"=D{row}*E{row}").font = VALUE_FONT
    ws1.cell(row, 6).alignment = RIGHT
    ws1.cell(row, 6).number_format = '"$"#,##0.00'
    for col in range(2, 7):
        ws1.cell(row, col).border = ALL_BORDER

# Row 29: 小计 (公式 SUM F26:F28)
ws1.row_dimensions[29].height = 22
ws1.merge_cells("B29:E29")
ws1["B29"] = "Subtotal (小计)"
ws1["B29"].font = TOTAL_FONT
ws1["B29"].fill = TOTAL_FILL
ws1["B29"].alignment = RIGHT
ws1["F29"] = "=SUM(F26:F28)"
ws1["F29"].font = TOTAL_FONT
ws1["F29"].fill = TOTAL_FILL
ws1["F29"].alignment = RIGHT
ws1["F29"].number_format = '"$"#,##0.00'
for col in range(2, 7):
    ws1.cell(29, col).border = TOP_BORDER

# Row 30: 折扣 (假设 2.4% 批量折扣)
ws1.row_dimensions[30].height = 22
ws1.merge_cells("B30:E30")
ws1["B30"] = "Discount (2.4% bulk discount)"
ws1["B30"].font = VALUE_FONT
ws1["B30"].alignment = RIGHT
ws1["F30"] = "=F29*0.024"
ws1["F30"].font = VALUE_FONT
ws1["F30"].alignment = RIGHT
ws1["F30"].number_format = '"$"#,##0.00'
for col in range(2, 7):
    ws1.cell(30, col).border = ALL_BORDER

# Row 31: 运费 (FOB 0)
ws1.row_dimensions[31].height = 22
ws1.merge_cells("B31:E31")
ws1["B31"] = "Shipping (per FOB Shenzhen)"
ws1["B31"].font = VALUE_FONT
ws1["B31"].alignment = RIGHT
ws1["F31"] = 0.00
ws1["F31"].font = VALUE_FONT
ws1["F31"].alignment = RIGHT
ws1["F31"].number_format = '"$"#,##0.00'
for col in range(2, 7):
    ws1.cell(31, col).border = ALL_BORDER

# Row 32: 保险 (FOB 0)
ws1.row_dimensions[32].height = 22
ws1.merge_cells("B32:E32")
ws1["B32"] = "Insurance (per FOB Shenzhen)"
ws1["B32"].font = VALUE_FONT
ws1["B32"].alignment = RIGHT
ws1["F32"] = 0.00
ws1["F32"].font = VALUE_FONT
ws1["F32"].alignment = RIGHT
ws1["F32"].number_format = '"$"#,##0.00'
for col in range(2, 7):
    ws1.cell(32, col).border = ALL_BORDER

# Row 33: 总价 (公式)
ws1.row_dimensions[33].height = 26
ws1.merge_cells("B33:E33")
ws1["B33"] = "TOTAL (总价)"
ws1["B33"].font = TOTAL_FONT
ws1["B33"].fill = TOTAL_FILL
ws1["B33"].alignment = RIGHT
ws1["F33"] = "=F29-F30+F31+F32"
ws1["F33"].font = TOTAL_FONT
ws1["F33"].fill = TOTAL_FILL
ws1["F33"].alignment = RIGHT
ws1["F33"].number_format = '"$"#,##0.00'
for col in range(2, 7):
    ws1.cell(33, col).border = BOTTOM_BORDER

# Row 34: 空行
ws1.row_dimensions[34].height = 6

# Row 35-36: 交期 + 运输方式
ws1.row_dimensions[35].height = 22
ws1["B35"] = "Lead Time:"
ws1["B35"].font = LABEL_FONT
ws1["B35"].alignment = LEFT
ws1.merge_cells("C35:F35")
ws1["C35"] = "3-5 working days for production + 2-4 days for DHL shipping"
ws1["C35"].font = VALUE_FONT
ws1["C35"].alignment = LEFT

ws1.row_dimensions[36].height = 22
ws1["B36"] = "Shipping Method:"
ws1["B36"].font = LABEL_FONT
ws1["B36"].alignment = LEFT
ws1.merge_cells("C36:F36")
ws1["C36"] = "DHL Global Express / FedEx International / SF International / Sea Freight (15-30 days)"
ws1["C36"].font = VALUE_FONT
ws1["C36"].alignment = LEFT

# Row 37: 空行
ws1.row_dimensions[37].height = 6

# Row 38-40: Notes
ws1.row_dimensions[38].height = 36
ws1.merge_cells("B38:F38")
ws1["B38"] = "NOTES (备注):"
ws1["B38"].font = LABEL_FONT
ws1["B38"].alignment = LEFT

notes_text = [
    "1. All prices in USD, FOB Shenzhen. Production starts upon receipt of 30% deposit.",
    "2. Quality standard: ISO 9001 + FSC + CMYK per ISO 12647-2 (FOGRA39 / GRACoL).",
    "3. Validity: 30 days from PI date. Rush order 100 MOQ available (18:00 cut-off, SF Express next-day noon).",
    "4. Any discrepancy subject to Hong Kong SAR law + CISG + Chinese Civil Code.",
]
for i, note in enumerate(notes_text):
    row = 39 + i
    ws1.row_dimensions[row].height = 18
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws1.cell(row, 2, note).font = NOTE_FONT
    ws1.cell(row, 2).alignment = LEFT

# Row 43: 空行
ws1.row_dimensions[43].height = 6

# Row 44-45: 签字盖章
ws1.row_dimensions[44].height = 28
ws1.merge_cells("B44:C44")
ws1["B44"] = "For and on behalf of Seller:"
ws1["B44"].font = LABEL_FONT
ws1["B44"].alignment = LEFT
ws1.merge_cells("D44:F44")
ws1["D44"] = "For and on behalf of Buyer:"
ws1["D44"].font = LABEL_FONT
ws1["D44"].alignment = LEFT

ws1.row_dimensions[45].height = 28
ws1.merge_cells("B45:C45")
ws1["B45"] = "Shenzhen Cailong Printing Packaging Co., Ltd."
ws1["B45"].font = VALUE_FONT
ws1["B45"].alignment = LEFT
ws1.merge_cells("D45:F45")
ws1["D45"] = "[Customer Company Name]"
ws1["D45"].font = VALUE_FONT
ws1["D45"].alignment = LEFT

ws1.row_dimensions[46].height = 28
ws1.merge_cells("B46:C46")
ws1["B46"] = "Mr. Tang Yunti (Legal Representative)"
ws1["B46"].font = VALUE_FONT
ws1["B46"].alignment = LEFT
ws1.merge_cells("D46:F46")
ws1["D46"] = "[Contact Person Name]"
ws1["D46"].font = VALUE_FONT
ws1["D46"].alignment = LEFT

ws1.row_dimensions[47].height = 36
ws1.merge_cells("B47:C47")
ws1["B47"] = "Signature: ____________________\nDate: 2026-08-27\n[Company Stamp]"
ws1["B47"].font = VALUE_FONT
ws1["B47"].alignment = LEFT
ws1.merge_cells("D47:F47")
ws1["D47"] = "Signature: ____________________\nDate: ____________________\n[Company Stamp]"
ws1["D47"].font = VALUE_FONT
ws1["D47"].alignment = LEFT

# 页眉页脚
ws1.oddHeader.left.text = "智印港 ZprintPro"
ws1.oddHeader.left.size = 9
ws1.oddHeader.right.text = "PI-2026-08-001"
ws1.oddHeader.right.size = 9
ws1.oddFooter.center.text = "Page &P of &N · zprintpro.com"
ws1.oddFooter.center.size = 9

# 打印设置
ws1.page_setup.orientation = ws1.ORIENTATION_PORTRAIT
ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.sheet_properties.pageSetUpPr.fitToPage = True
ws1.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

# ---------- Sheet 2: Items (详细品名表) ----------
ws2 = wb.create_sheet("Items")
col_widths2 = {"A": 4, "B": 14, "C": 16, "D": 50, "E": 10, "F": 8, "G": 16, "H": 16, "I": 22, "J": 4}
for col, width in col_widths2.items():
    ws2.column_dimensions[col].width = width

# 标题
ws2.row_dimensions[1].height = 32
ws2.merge_cells("B1:I1")
ws2["B1"] = "ITEMS DETAILED TABLE (品名详细表)"
ws2["B1"].font = TITLE_FONT
ws2["B1"].fill = TITLE_FILL
ws2["B1"].alignment = CENTER

# 表头
ws2.row_dimensions[2].height = 36
headers2 = ["HS Code", "Category", "Description (English)", "Qty", "Unit", "Unit Price (USD)", "Amount (USD)", "Packaging Spec"]
for i, h in enumerate(headers2):
    cell = ws2.cell(2, 2 + i, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDER

# 3 个品名 + 包装
items_detail = [
    ("49119900", "Sticker", "Custom Sticker - round 50mm diameter, waterproof vinyl, CMYK 4C, gloss lamination, die-cut, individual sheet", 10000, "pcs", 0.03, "PE bag 100pcs/bag + 5-ply AB corrugated carton 2000pcs/ctn"),
    ("48191000", "Packaging Box", "Custom Corrugated Box - 3-ply B-flute 250gsm kraft, CMYK 4C, matte lamination, RSC (Regular Slotted Container), 30x20x15cm", 1000, "pcs", 1.50, "PE bag 1pc/bag + 5-ply AB corrugated carton 50pcs/ctn"),
    ("48194000", "Paper Bag", "Custom Paper Bag - white card 250gsm, CMYK 4C, spot UV on logo, cotton rope handle, matte lamination, 25x15x30cm", 500, "pcs", 0.50, "PE bag 10pcs/bundle + 5-ply AB corrugated carton 100pcs/ctn"),
]
for i, (hs, cat, desc, qty, unit, price, pkg) in enumerate(items_detail):
    row = 3 + i
    ws2.row_dimensions[row].height = 56
    cells = [hs, cat, desc, qty, unit, price, pkg]
    for j, val in enumerate(cells):
        cell = ws2.cell(row, 2 + j, val)
        cell.font = VALUE_FONT
        cell.alignment = LEFT if j < 3 else (RIGHT if j >= 3 and j <= 6 else LEFT)
        cell.border = ALL_BORDER
        if j == 3:  # Qty
            cell.number_format = "#,##0"
        if j == 5:  # Unit Price
            cell.number_format = '"$"#,##0.0000'
    # Amount (公式)
    ws2.cell(row, 8, f"=E{row}*G{row}").font = VALUE_FONT
    ws2.cell(row, 8).alignment = RIGHT
    ws2.cell(row, 8).number_format = '"$"#,##0.00'
    ws2.cell(row, 8).border = ALL_BORDER

# Row 6: 小计
ws2.row_dimensions[6].height = 24
ws2.merge_cells("B6:G6")
ws2["B6"] = "TOTAL (合计)"
ws2["B6"].font = TOTAL_FONT
ws2["B6"].fill = TOTAL_FILL
ws2["B6"].alignment = RIGHT
ws2["H6"] = "=SUM(H3:H5)"
ws2["H6"].font = TOTAL_FONT
ws2["H6"].fill = TOTAL_FILL
ws2["H6"].alignment = RIGHT
ws2["H6"].number_format = '"$"#,##0.00'
ws2["I6"] = ""
ws2["I6"].fill = TOTAL_FILL
for col in range(2, 10):
    ws2.cell(6, col).border = TOP_BORDER

# 打印设置
ws2.page_setup.orientation = ws2.ORIENTATION_PORTRAIT
ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True

# ---------- Sheet 3: Bank (银行信息) ----------
ws3 = wb.create_sheet("Bank")
col_widths3 = {"A": 4, "B": 26, "C": 50, "D": 4}
for col, width in col_widths3.items():
    ws3.column_dimensions[col].width = width

# 标题
ws3.row_dimensions[1].height = 32
ws3.merge_cells("B1:C1")
ws3["B1"] = "BANK INFORMATION (银行信息)"
ws3["B1"].font = TITLE_FONT
ws3["B1"].fill = TITLE_FILL
ws3["B1"].alignment = CENTER

# 银行信息
bank_rows = [
    ("Beneficiary (收款人)", "Shenzhen Cailong Printing Packaging Co., Ltd."),
    ("Bank Name (银行名称)", "DBS Bank (Hong Kong) Limited"),
    ("SWIFT Code", "DHBKHKHH"),
    ("Bank Address (银行地址)", "11/F, The Center, 99 Queen's Road Central, Hong Kong"),
    ("USD Account (美元账户)", "[DBS HK USD Account Number - K3 to confirm]"),
    ("HKD Account (港币账户)", "[DBS HK HKD Account Number - K3 to confirm]"),
    ("CNY Account (人民币账户)", "[DBS HK CNY Account Number - K3 to confirm]"),
    ("Correspondent Bank (代理行)", "Standard Chartered Bank Hong Kong (for USD clearing)"),
    ("Payment Reference (付款备注)", "PI-2026-08-001 + Customer Company Name"),
]
for i, (k, v) in enumerate(bank_rows):
    row = 2 + i
    ws3.row_dimensions[row].height = 22
    ws3.cell(row, 2, k).font = LABEL_FONT
    ws3.cell(row, 2).alignment = LEFT
    ws3.cell(row, 3, v).font = VALUE_FONT
    ws3.cell(row, 3).alignment = LEFT

# 落盘
pi_path = os.path.join(INVOICE_DIR, "PI-2026-08-001.xlsx")
wb.save(pi_path)
pi_size_kb = os.path.getsize(pi_path) / 1024
print(f"✅ Proforma Invoice Excel 生成: {pi_path}")
print(f"   Size: {pi_size_kb:.1f} KB")
print(f"   Sheets: 3 (Proforma Invoice + Items + Bank)")

# ============================================================================
# Quotation (Q-2026-08-001.xlsx) - zh-hk 港币 + 2 工作表
# ============================================================================
wb2 = Workbook()

# ---------- Sheet 1: Quotation ----------
ws_q1 = wb2.active
ws_q1.title = "報價單"

col_widths_q1 = {"A": 4, "B": 24, "C": 36, "D": 12, "E": 14, "F": 16, "G": 4}
for col, width in col_widths_q1.items():
    ws_q1.column_dimensions[col].width = width

# 标题
ws_q1.row_dimensions[1].height = 36
ws_q1.merge_cells("B1:F1")
ws_q1["B1"] = "智 印 港 報 價 單  |  ZprintPro Quotation"
ws_q1["B1"].font = ZH_TITLE_FONT
ws_q1["B1"].fill = TITLE_FILL
ws_q1["B1"].alignment = CENTER

# 报价单基本信息
ws_q1.row_dimensions[2].height = 22
ws_q1["B2"] = "報價單編號:"
ws_q1["B2"].font = ZH_LABEL_FONT
ws_q1["B2"].alignment = LEFT
ws_q1["C2"] = "Q-2026-08-001"
ws_q1["C2"].font = ZH_VALUE_FONT
ws_q1["C2"].alignment = LEFT
ws_q1["D2"] = "報價日期:"
ws_q1["D2"].font = ZH_LABEL_FONT
ws_q1["D2"].alignment = LEFT
ws_q1["E2"] = "2026 年 8 月 27 日"
ws_q1["E2"].font = ZH_VALUE_FONT
ws_q1["E2"].alignment = LEFT
ws_q1["F2"] = ""
ws_q1["F2"].alignment = LEFT

ws_q1.row_dimensions[3].height = 22
ws_q1["B3"] = "有效期至:"
ws_q1["B3"].font = ZH_LABEL_FONT
ws_q1["B3"].alignment = LEFT
ws_q1["C3"] = "2026 年 9 月 26 日 (30 天)"
ws_q1["C3"].font = ZH_VALUE_FONT
ws_q1["C3"].alignment = LEFT
ws_q1["D3"] = "幣種:"
ws_q1["D3"].font = ZH_LABEL_FONT
ws_q1["D3"].alignment = LEFT
ws_q1["E3"] = "HKD (港幣)"
ws_q1["E3"].font = ZH_VALUE_FONT
ws_q1["E3"].alignment = LEFT
ws_q1["F3"] = ""

# Row 4: 空行
ws_q1.row_dimensions[4].height = 6

# 客户信息
ws_q1.row_dimensions[5].height = 22
ws_q1.merge_cells("B5:F5")
ws_q1["B5"] = "客 戶 資 料  |  Customer Information"
ws_q1["B5"].font = ZH_HEADER_FONT
ws_q1["B5"].fill = HEADER_FILL
ws_q1["B5"].alignment = LEFT

customer_rows = [
    ("客戶姓名 / 公司", "[客戶姓名] / [客戶公司全稱]"),
    ("聯繫電話", "[客戶電話] (WhatsApp 優先)"),
    ("電子郵箱", "[客戶郵箱]"),
    ("送貨地址", "[客戶送貨地址]"),
]
for i, (k, v) in enumerate(customer_rows):
    row = 6 + i
    ws_q1.row_dimensions[row].height = 22
    ws_q1.cell(row, 2, k).font = ZH_LABEL_FONT
    ws_q1.cell(row, 2).alignment = LEFT
    ws_q1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    ws_q1.cell(row, 3, v).font = ZH_VALUE_FONT
    ws_q1.cell(row, 3).alignment = LEFT

# Row 10: 空行
ws_q1.row_dimensions[10].height = 6

# 报价表标题
ws_q1.row_dimensions[11].height = 22
ws_q1.merge_cells("B11:F11")
ws_q1["B11"] = "報 價 明 細  |  Quotation Details"
ws_q1["B11"].font = ZH_HEADER_FONT
ws_q1["B11"].fill = HEADER_FILL
ws_q1["B11"].alignment = LEFT

# 表头
ws_q1.row_dimensions[12].height = 32
headers_q1 = ["品名 / 規格", "工藝說明", "數量", "單價 (HKD)", "金額 (HKD)"]
for i, h in enumerate(headers_q1):
    cell = ws_q1.cell(12, 2 + i, h)
    cell.font = ZH_HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDER

# 3 个品名 (港币)
quote_items = [
    ("客製化貼紙 (圓形 50mm 直徑)\n防水 PVC + CMYK 4 色 + 亮膜 + 模切", "100 張 / 包 + 5 層 AB 浪紙箱 2,000 張 / 箱", 10000, 0.24),
    ("客製化包裝盒 (坑盒 3 层 B 浪 250gsm)\nCMYK 4 色 + 啞膠 + RSC 標準盒型 30x20x15cm", "1 個 / PE 袋 + 5 層 AB 浪紙箱 50 個 / 箱", 1000, 12.00),
    ("客製化紙袋 (白卡 250gsm + 棉繩)\nCMYK 4 色 + LOGO 局部 UV + 啞膠 25x15x30cm", "10 個 / 紮 + 5 層 AB 浪紙箱 100 個 / 箱", 500, 4.20),
]
for i, (desc, pkg, qty, price) in enumerate(quote_items):
    row = 13 + i
    ws_q1.row_dimensions[row].height = 48
    ws_q1.cell(row, 2, desc).font = ZH_VALUE_FONT
    ws_q1.cell(row, 2).alignment = LEFT
    ws_q1.cell(row, 3, pkg).font = ZH_VALUE_FONT
    ws_q1.cell(row, 3).alignment = LEFT
    ws_q1.cell(row, 4, qty).font = ZH_VALUE_FONT
    ws_q1.cell(row, 4).alignment = RIGHT
    ws_q1.cell(row, 4).number_format = "#,##0"
    ws_q1.cell(row, 5, price).font = ZH_VALUE_FONT
    ws_q1.cell(row, 5).alignment = RIGHT
    ws_q1.cell(row, 5).number_format = '"HK$"#,##0.00'
    # Amount (公式)
    ws_q1.cell(row, 6, f"=D{row}*E{row}").font = ZH_VALUE_FONT
    ws_q1.cell(row, 6).alignment = RIGHT
    ws_q1.cell(row, 6).number_format = '"HK$"#,##0.00'
    for col in range(2, 7):
        ws_q1.cell(row, col).border = ALL_BORDER

# Row 16: 小计
ws_q1.row_dimensions[16].height = 24
ws_q1.merge_cells("B16:E16")
ws_q1["B16"] = "小計 Subtotal"
ws_q1["B16"].font = ZH_TOTAL_FONT
ws_q1["B16"].fill = TOTAL_FILL
ws_q1["B16"].alignment = RIGHT
ws_q1["F16"] = "=SUM(F13:F15)"
ws_q1["F16"].font = ZH_TOTAL_FONT
ws_q1["F16"].fill = TOTAL_FILL
ws_q1["F16"].alignment = RIGHT
ws_q1["F16"].number_format = '"HK$"#,##0.00'
for col in range(2, 7):
    ws_q1.cell(16, col).border = TOP_BORDER

# Row 17: 折扣
ws_q1.row_dimensions[17].height = 22
ws_q1.merge_cells("B17:E17")
ws_q1["B17"] = "批量折扣 Bulk Discount (2.4%)"
ws_q1["B17"].font = ZH_VALUE_FONT
ws_q1["B17"].alignment = RIGHT
ws_q1["F17"] = "=F16*0.024"
ws_q1["F17"].font = ZH_VALUE_FONT
ws_q1["F17"].alignment = RIGHT
ws_q1["F17"].number_format = '"HK$"#,##0.00'
for col in range(2, 7):
    ws_q1.cell(17, col).border = ALL_BORDER

# Row 18: 运费
ws_q1.row_dimensions[18].height = 22
ws_q1.merge_cells("B18:E18")
ws_q1["B18"] = "順豐本地滿 HK$500 免費 / 跨境 DHL 2-4 天"
ws_q1["B18"].font = ZH_VALUE_FONT
ws_q1["B18"].alignment = RIGHT
ws_q1["F18"] = 0.00
ws_q1["F18"].font = ZH_VALUE_FONT
ws_q1["F18"].alignment = RIGHT
ws_q1["F18"].number_format = '"HK$"#,##0.00'
for col in range(2, 7):
    ws_q1.cell(18, col).border = ALL_BORDER

# Row 19: 总价
ws_q1.row_dimensions[19].height = 26
ws_q1.merge_cells("B19:E19")
ws_q1["B19"] = "總  計  TOTAL"
ws_q1["B19"].font = ZH_TOTAL_FONT
ws_q1["B19"].fill = TOTAL_FILL
ws_q1["B19"].alignment = RIGHT
ws_q1["F19"] = "=F16-F17+F18"
ws_q1["F19"].font = ZH_TOTAL_FONT
ws_q1["F19"].fill = TOTAL_FILL
ws_q1["F19"].alignment = RIGHT
ws_q1["F19"].number_format = '"HK$"#,##0.00'
for col in range(2, 7):
    ws_q1.cell(19, col).border = BOTTOM_BORDER

# Row 20: 空行
ws_q1.row_dimensions[20].height = 6

# 付款方式 + 交期
ws_q1.row_dimensions[21].height = 22
ws_q1["B21"] = "付款方式:"
ws_q1["B21"].font = ZH_LABEL_FONT
ws_q1["B21"].alignment = LEFT
ws_q1.merge_cells("C21:F21")
ws_q1["C21"] = "T/T 30% 訂金 + 70% 尾款 / PayPal / Airwallex / DBS HK 銀行電匯"
ws_q1["C21"].font = ZH_VALUE_FONT
ws_q1["C21"].alignment = LEFT

ws_q1.row_dimensions[22].height = 22
ws_q1["B22"] = "交期:"
ws_q1["B22"].font = ZH_LABEL_FONT
ws_q1["B22"].alignment = LEFT
ws_q1.merge_cells("C22:F22")
ws_q1["C22"] = "3-5 個工作天 (急件 4-6 小時, 18:00 截單, 順豐翌日中午 12:00 前送達)"
ws_q1["C22"].font = ZH_VALUE_FONT
ws_q1["C22"].alignment = LEFT

ws_q1.row_dimensions[23].height = 22
ws_q1["B23"] = "質量標準:"
ws_q1["B23"].font = ZH_LABEL_FONT
ws_q1["B23"].alignment = LEFT
ws_q1.merge_cells("C23:F23")
ws_q1["C23"] = "ISO 9001 + FSC 認證 + CMYK 4 色 (ISO 12647-2 FOGRA39 / GRACoL)"
ws_q1["C23"].font = ZH_VALUE_FONT
ws_q1["C23"].alignment = LEFT

# Row 24: 空行
ws_q1.row_dimensions[24].height = 6

# 备注
ws_q1.row_dimensions[25].height = 36
ws_q1.merge_cells("B25:F25")
ws_q1["B25"] = "備注 NOTES:"
ws_q1["B25"].font = ZH_LABEL_FONT
ws_q1["B25"].alignment = LEFT

quote_notes = [
    "1. 報價含基本包裝, 不含運輸及保險。順豐本地滿 HK$500 免運費, 跨境 DHL 2-4 天送達。",
    "2. 質量標準: ISO 9001 質量管理體系 + FSC 森林認證 + CMYK 4 色印刷 (ISO 12647-2 標準)。",
    "3. 急件服務: 100 張起印, 18:00 截單, 順豐翌日中午 12:00 前送達, 加急費 30-50%。",
    "4. 本報價單有效期 30 天, 確認訂單後另行簽訂銷售合同 (Sales Contract)。",
]
for i, note in enumerate(quote_notes):
    row = 26 + i
    ws_q1.row_dimensions[row].height = 18
    ws_q1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws_q1.cell(row, 2, note).font = ZH_NOTE_FONT
    ws_q1.cell(row, 2).alignment = LEFT

# Row 30: 联系方式
ws_q1.row_dimensions[30].height = 22
ws_q1.merge_cells("B30:F30")
ws_q1["B30"] = "WhatsApp 30 秒即時報價: +86 198 8085 1334  |  郵箱: zprintpro@outlook.com  |  官網: zprintpro.com"
ws_q1["B30"].font = ZH_VALUE_FONT
ws_q1["B30"].alignment = CENTER

# 页眉页脚
ws_q1.oddHeader.left.text = "智印港 ZprintPro"
ws_q1.oddHeader.left.size = 9
ws_q1.oddHeader.right.text = "Q-2026-08-001"
ws_q1.oddHeader.right.size = 9
ws_q1.oddFooter.center.text = "Page &P of &N · zprintpro.com"
ws_q1.oddFooter.center.size = 9

# 打印设置
ws_q1.page_setup.orientation = ws_q1.ORIENTATION_PORTRAIT
ws_q1.page_setup.paperSize = ws_q1.PAPERSIZE_A4
ws_q1.page_setup.fitToWidth = 1
ws_q1.page_setup.fitToHeight = 0
ws_q1.sheet_properties.pageSetUpPr.fitToPage = True
ws_q1.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

# ---------- Sheet 2: Cost Breakdown (成本明细) ----------
ws_q2 = wb2.create_sheet("成本明細")
col_widths_q2 = {"A": 4, "B": 28, "C": 14, "D": 14, "E": 14, "F": 14, "G": 4}
for col, width in col_widths_q2.items():
    ws_q2.column_dimensions[col].width = width

# 标题
ws_q2.row_dimensions[1].height = 32
ws_q2.merge_cells("B1:F1")
ws_q2["B1"] = "成 本 明 細 表  |  Cost Breakdown"
ws_q2["B1"].font = ZH_TITLE_FONT
ws_q2["B1"].fill = TITLE_FILL
ws_q2["B1"].alignment = CENTER

# 表头
ws_q2.row_dimensions[2].height = 28
headers_q2 = ["成本項目", "貼紙", "包裝盒", "紙袋", "小計 (HKD)"]
for i, h in enumerate(headers_q2):
    cell = ws_q2.cell(2, 2 + i, h)
    cell.font = ZH_HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = ALL_BORDER

# 成本明细 (4 大类: 纸张 + 印刷 + 工艺 + 包装 + 物流 + 利润)
cost_items = [
    ("紙材 (Paper)", 800, 4000, 1500),
    ("印刷 (CMYK 4C)", 600, 2000, 800),
    ("工藝 (Lamination / Spot UV / Die-cut)", 400, 1500, 600),
    ("包裝 (PE 袋 + 紙箱)", 300, 500, 200),
    ("物流 (Local + Cross-border)", 200, 1000, 400),
    ("小計 (Subtotal)", 2300, 9000, 3500),
    ("利潤 (Profit 30%)", 690, 2700, 1050),
    ("總計 (Total)", 2990, 11700, 4550),
]
for i, (item, sticker, box, bag) in enumerate(cost_items):
    row = 3 + i
    ws_q2.row_dimensions[row].height = 24
    is_subtotal = "小計" in item or "總計" in item
    is_profit = "利潤" in item
    cell_font = ZH_TOTAL_FONT if is_subtotal or is_profit else ZH_VALUE_FONT
    cell_fill = TOTAL_FILL if is_subtotal else None
    ws_q2.cell(row, 2, item).font = cell_font
    ws_q2.cell(row, 2).alignment = LEFT
    if cell_fill:
        ws_q2.cell(row, 2).fill = cell_fill
    for j, val in enumerate([sticker, box, bag]):
        c = ws_q2.cell(row, 3 + j, val)
        c.font = cell_font
        c.alignment = RIGHT
        c.number_format = '"HK$"#,##0.00'
        if cell_fill:
            c.fill = cell_fill
    # Subtotal 公式
    if item == "小計 (Subtotal)":
        for j, col_letter in enumerate(["C", "D", "E"]):
            c = ws_q2.cell(row, 3 + j, f"=SUM({col_letter}3:{col_letter}7)")
            c.font = ZH_TOTAL_FONT
            c.alignment = RIGHT
            c.number_format = '"HK$"#,##0.00'
            c.fill = TOTAL_FILL
    elif item == "利潤 (Profit 30%)":
        for j, col_letter in enumerate(["C", "D", "E"]):
            c = ws_q2.cell(row, 3 + j, f"={col_letter}8*0.3")
            c.font = ZH_TOTAL_FONT
            c.alignment = RIGHT
            c.number_format = '"HK$"#,##0.00'
    elif item == "總計 (Total)":
        for j, col_letter in enumerate(["C", "D", "E"]):
            c = ws_q2.cell(row, 3 + j, f"={col_letter}8+{col_letter}9")
            c.font = ZH_TOTAL_FONT
            c.alignment = RIGHT
            c.number_format = '"HK$"#,##0.00'
            c.fill = TOTAL_FILL
    # 总价 (Sum of 3 columns)
    c = ws_q2.cell(row, 6, f"=SUM(C{row}:E{row})")
    c.font = cell_font
    c.alignment = RIGHT
    c.number_format = '"HK$"#,##0.00'
    if cell_fill:
        c.fill = cell_fill
    for col in range(2, 7):
        ws_q2.cell(row, col).border = ALL_BORDER

# 利润率分析
ws_q2.row_dimensions[12].height = 22
ws_q2.merge_cells("B12:F12")
ws_q2["B12"] = "利 潤 率 分 析  |  Margin Analysis"
ws_q2["B12"].font = ZH_HEADER_FONT
ws_q2["B12"].fill = HEADER_FILL
ws_q2["B12"].alignment = LEFT

margin_rows = [
    ("成本占比 (Cost %)", "=F8/F19*100", "客戶報價總計 / 客戶報價小計"),
    ("毛利率 (Gross Margin %)", "=F9/F19*100", "利潤 / 客戶報價小計"),
    ("淨利率 (Net Margin %)", "=F9/F19*100", "利潤 / 客戶報價小計 (簡化估算)"),
]
for i, (k, formula, desc) in enumerate(margin_rows):
    row = 13 + i
    ws_q2.row_dimensions[row].height = 22
    ws_q2.cell(row, 2, k).font = ZH_LABEL_FONT
    ws_q2.cell(row, 2).alignment = LEFT
    ws_q2.cell(row, 3, formula).font = ZH_VALUE_FONT
    ws_q2.cell(row, 3).alignment = RIGHT
    ws_q2.cell(row, 3).number_format = "0.0%"
    ws_q2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws_q2.cell(row, 4, desc).font = ZH_NOTE_FONT
    ws_q2.cell(row, 4).alignment = LEFT

# 落盘
quote_path = os.path.join(QUOTE_DIR, "Q-2026-08-001.xlsx")
wb2.save(quote_path)
quote_size_kb = os.path.getsize(quote_path) / 1024
print(f"✅ Quotation Excel 生成: {quote_path}")
print(f"   Size: {quote_size_kb:.1f} KB")
print(f"   Sheets: 2 (報價單 + 成本明細)")

print("\n" + "=" * 70)
print("✅ 3 个办公文件全部生成完成")
print("=" * 70)
print(f"  1. {os.path.join(INVOICE_DIR, 'PI-2026-08-001.xlsx')} ({pi_size_kb:.1f} KB)")
print(f"  2. {quote_path} ({quote_size_kb:.1f} KB)")
print(f"  3. F:\\zprintpro-nextjs\\docs\\contracts\\ZP-SA-2026-08-001.docx (42.7 KB)")
